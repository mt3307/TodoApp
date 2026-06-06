import os

from fastapi import FastAPI
from fastapi.responses  import FileResponse
from sqlalchemy import create_engine
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import date

app = FastAPI()

engine = create_engine(
    "mysql+pymysql://root:Mt33070402%40@localhost/todoapp"
)

#テーブル情報の取得
@app.get("/todos/{user_id}")
def get_user_todos(user_id: int):
    with engine.connect() as conn:
        result = conn.execute(
            text("""
                 SELECT * from todos
                 WHERE user_id = :user_id
                 ORDER BY task_date
                 """),
                 {"user_id":user_id}
                 )
        
        rows = result.fetchall()

        return [
            dict(row._mapping)
            for row in rows
        ]

#Excelを生成
@app.get("/excel/{user_id}")
def export_excel(user_id: int):
    with engine.connect() as conn:
        result = conn.execute(
            text("""
                 SELECT * FROM todos 
                 WHERE user_id = :user_id
                 ORDER BY task_date
                 """),
                 {"user_id": user_id}
        )

        todos = result.fetchall()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Todo"
    ws.append(
        [
            "ID",
            "タスク",
            "期限",
            "完了"
        ]
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFCCCC",
        end_color="FFCCCC"  
    ) 

    today = date.today()
    print("today=", today)
    for todo in todos:

        print(todo.task_date, type(todo.task_date))
        print(todo.completed, type(todo.completed))

        status = "完了" if todo.completed else "未完了"
        ws.append(
            [
                todo.id,
                todo.task,
                str(todo.task_date),
                status
            ]
        )

        current_row = ws.max_row
        if (todo.task_date < today and not todo.completed) :
            for cell in ws[current_row]:
                cell.fill = red_fill
    
    total_count = len(todos)
    completed_count = sum(
        1 for todo in todos
        if todo.completed
    )

    incompleted_count = (
        total_count - completed_count
    )

    ws.append([])
    ws.append(["----- 集計情報 -----"])
    ws.append(["総件数", total_count])
    ws.append(["完了件数", completed_count])
    ws.append(["未完了件数", incompleted_count])

    if total_count == 0:
        completion_rate = 0
    else:
        completion_rate = round(
            completed_count / total_count * 100, 1
        )
    ws.append(["完了率", f"{completion_rate}%"])

    os.makedirs(
        "__Excel",
        exist_ok=True
    )
    filename = f"__Excel/todo_{user_id}.xlsx"
    wb.save(filename)

    return FileResponse(
        filename,
        filename="todo.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )